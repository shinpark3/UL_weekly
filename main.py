import os
import argparse
import subprocess
import sys
import datetime as dt
import pandas as pd
import numpy as np
import dateutil.relativedelta
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from pyspark.sql import SparkSession

reload(sys)
sys.setdefaultencoding('utf8')

SPARK = SparkSession \
    .builder \
    .enableHiveSupport() \
    .getOrCreate()
SPARK.sql('use shopee')


def get_MY():
    query = '''
        SELECT distinct(c.sku_id), c.grass_date, c.real_sales
        FROM shopee_bi_replenishment_data_pool_core c
        JOIN shopee_bi_replenishment_data_pool_ext_replenishment_info e  
        ON c.sku_id=e.sku_id
        WHERE e.sourcing_status = 0
        AND c.grass_date >= current_date - interval '31' day
        AND c.grass_region ='MY'
        AND c.shopid = 26152357
        GROUP BY c.sku_id, c.grass_date, c.real_sales
        ORDER BY c.grass_date ASC 
        '''
    print(query)
    return SPARK.sql(query)


def get_PH():
    query = '''
        SELECT distinct(c.sku_id), c.grass_date, c.real_sales
        FROM shopee_bi_replenishment_data_pool_core c
        JOIN shopee_bi_replenishment_data_pool_ext_replenishment_info e  
        ON c.sku_id=e.sku_id
        WHERE e.sourcing_status = 0
        AND c.grass_date >= current_date - interval '31' day
        AND c.grass_region ='PH'
        AND c.shopid in (40492624, 40867978, 40495882)
        GROUP BY c.sku_id, c.grass_date, c.real_sales
        ORDER BY c.grass_date ASC 
        '''
    print(query)
    return SPARK.sql(query)


def get_ID():
    query = '''
        SELECT distinct(c.sku_id), c.grass_date, c.real_sales
        FROM shopee_bi_replenishment_data_pool_core c
        JOIN shopee_bi_replenishment_data_pool_ext_replenishment_info e  
        ON c.sku_id=e.sku_id
        WHERE e.sourcing_status = 0
        AND c.grass_date >= current_date - interval '31' day
        AND c.grass_region ='ID'
        AND c.shopid = 14318452
        GROUP BY c.sku_id, c.grass_date, c.real_sales
        ORDER BY c.grass_date ASC 
        '''
    print(query)
    return SPARK.sql(query)


def main():
    my_data = get_MY()
    ph_data = get_PH()
    id_data = get_ID()
    df_list = [my_data, ph_data, id_data]
    country_list = ['MY', 'PH', 'ID']
    for df, country in zip(df_list, country_list):
        df['real_sales'] = np.where(df['real_sales'] != 'x', 0, df['real_sales'])
        df['real_sales'].replace('x', 1, inplace=True)
        df['sku_id'] = df['sku_id'].str.split('_').str[0]
        df['sku_id'] = pd.to_numeric(df['sku_id'])
    my_pivot = my_data.pivot_table(values='real_sales', columns='grass_date', index='sku_id').reset_index()
    ph_pivot = ph_data.pivot_table(values='real_sales', columns='grass_date', index='sku_id').reset_index()
    id_pivot = id_data.pivot_table(values='real_sales', columns='grass_date', index='sku_id').reset_index()

    filename = './UL_ATP_summary_template_v2.xlsx'
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = load_workbook(filename)
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    pivot_list = [my_pivot, ph_pivot, id_pivot]
    for df, country in zip(pivot_list, country_list):
        sheet_name = '{}_Pivot'.format(country)
        df.to_excel(writer, sheet_name, index=False, header=True, startrow=0, startcol=0)
    output_file_name = './{}_UL_ATP_Summary_{}.xlsx'.format(country, dt.date.today().strftime('%Y_%m_%d'))
    writer.book.save(output_file_name)


if __name__ == '__main__':
    main()
